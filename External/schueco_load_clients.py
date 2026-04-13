#!/usr/bin/env python3
"""
Schüco Clients Loader via OData

Loads Партнер + Контрагент + КонтактныеЛицаПартнеров + БанковскиеСчетаКонтрагентов
from a sales-team Excel workbook (Clients.xlsx) into BAS UT 3.5 via the OData REST API.

Sibling of the Schueco_ЗагрузкаКонтрагентов.epf external data processor.
Reads the 32-column flat layout (Name / Address components / EDRPOU / IPN / Bank / Contacts)
and applies the same dedup logic.

Multi-sheet XLSX: use --sheet to pick a sheet (default = first sheet).

Addresses are written with structured JSON in Значение (city, street, houseNumber, etc.)
matching the BSL sibling's ДобавитьАдрес format. Phone/email КИ rows still use only
Представление — the structured Значение/ЗначенияПолей for those types requires
server-side УправлениеКонтактнойИнформацією which is not exposed via OData.

Usage:
    python3 schueco_load_clients.py --excel Clients.xlsx --sheet "Ivan Tereshkevych"
    python3 schueco_load_clients.py --excel Clients.xlsx --manager "Адміненко" --update
    python3 schueco_load_clients.py --excel Clients.xlsx --dry-run
"""

import argparse
import base64
import json
import re
import sys
from typing import Optional

import openpyxl
import requests

# === CONFIGURATION ===
BASE_URL = "http://10.1.5.109/ut_demo/odata/standard.odata"
# Cyrillic username needs UTF-8 encoding for Basic auth (latin-1 fails)
_AUTH_STR = base64.b64encode("Адміненко:".encode("utf-8")).decode("ascii")
HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
    "Authorization": f"Basic {_AUTH_STR}",
}

DRY_RUN = False

# === HELPERS ===

def odata_get(entity: str, params: str = "") -> dict:
    """GET from OData endpoint."""
    url = f"{BASE_URL}/{entity}"
    if params:
        url += f"?{params}"
    resp = requests.get(url, headers=HEADERS, timeout=30)
    resp.raise_for_status()
    return resp.json()


def odata_post(entity: str, payload: dict) -> dict:
    """POST to OData endpoint. Returns created object or raises."""
    if DRY_RUN:
        print(f"  [DRY] POST {entity} -> {payload.get('Description', '?')}")
        # Return a fake object so the caller can chain without crashing
        return {"Ref_Key": "00000000-0000-0000-0000-000000000000", **payload}
    url = f"{BASE_URL}/{entity}"
    resp = requests.post(url, headers=HEADERS, json=payload, timeout=30)
    if resp.status_code >= 400:
        try:
            err = resp.json()
            msg = err.get("odata.error", {}).get("message", {}).get("value", resp.text)
        except Exception:
            msg = resp.text
        raise RuntimeError(f"OData POST {entity} failed ({resp.status_code}): {msg}")
    return resp.json()


def odata_patch(entity: str, ref_key: str, payload: dict) -> dict:
    """PATCH existing object by Ref_Key."""
    if DRY_RUN:
        print(f"  [DRY] PATCH {entity}({ref_key[:8]}...) -> {list(payload.keys())}")
        return {"Ref_Key": ref_key, **payload}
    url = f"{BASE_URL}/{entity}(guid'{ref_key}')"
    resp = requests.patch(url, headers=HEADERS, json=payload, timeout=30)
    if resp.status_code >= 400:
        try:
            err = resp.json()
            msg = err.get("odata.error", {}).get("message", {}).get("value", resp.text)
        except Exception:
            msg = resp.text
        raise RuntimeError(f"OData PATCH {entity} failed ({resp.status_code}): {msg}")
    return resp.json()


def safe_str(val) -> str:
    """Convert cell value to stripped string, replacing non-breaking spaces."""
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip()


def safe_digits(val) -> str:
    """Extract digits only from a cell value."""
    s = safe_str(val)
    return "".join(ch for ch in s if ch.isdigit())


def odata_quote(s: str) -> str:
    """Escape a string for embedding in an OData $filter literal."""
    return s.replace("'", "''")


# === ENUM / REF LOOKUPS ===

def detect_ur_fiz(name: str) -> str:
    """Return the ЮрФизЛицо enum string value based on the name prefix.
    Mirrors ОпределитьЮрФизЛицо() in the BSL sibling."""
    n = name.strip().upper()
    if n.startswith(("ФОП ", "ФОП.", "СПД ", "СПД.")):
        return "ИндивидуальныйПредприниматель"
    if n.startswith((
        "ФІЗ. ОСОБА", "ФИЗ. ЛИЦО", "ФІЗИЧНА ОСОБА", "ФИЗИЧЕСКОЕ ЛИЦО",
        "ЧАСТНОЕ ЛИЦО", "ПРИВАТНА ОСОБА",
    )):
        return "ФизЛицо"
    return "ЮрЛицо"


def partner_ur_fiz(ur_fiz: str) -> str:
    """Map ЮрФизЛицо → КомпанияЧастноеЛицо (Партнер enum)."""
    return "ЧастноеЛицо" if ur_fiz == "ФизЛицо" else "Компания"


class RefCache:
    """Lazy cache for Ref_Keys of ВидыКонтактнойИнформации and other lookup catalogs."""

    def __init__(self):
        self._ki_types: dict[str, str] = {}
        self._banks: dict[str, Optional[str]] = {}

    def ki_type(self, name: str) -> Optional[str]:
        if name in self._ki_types:
            return self._ki_types[name]
        data = odata_get(
            "Catalog_ВидыКонтактнойИнформации",
            f"$filter=PredefinedDataName eq '{name}'&$select=Ref_Key&$top=1",
        )
        items = data.get("value", [])
        key = items[0]["Ref_Key"] if items else None
        self._ki_types[name] = key
        return key

    def bank_by_mfo(self, mfo: str) -> Optional[str]:
        if mfo in self._banks:
            return self._banks[mfo]
        try:
            data = odata_get(
                "Catalog_КлассификаторБанков",
                f"$filter=Code eq '{odata_quote(mfo)}'&$select=Ref_Key&$top=1",
            )
            items = data.get("value", [])
            key = items[0]["Ref_Key"] if items else None
        except Exception:
            key = None
        self._banks[mfo] = key
        return key


# === PHASE 1: READ EXCEL ===
# Clients_Flat.xlsx layout (35 columns A–AI):
#   A  (0)  Manager          — sales manager name
#   B  (1)  Company Name     — full legal name (НаименованиеПолное)
#   C  (2)  Short Name       — short display name (Наименование, up to 100 chars)
#   D  (3)  Company Address  — presentation text
#   E  (4)  Індекс           — ZIP code
#   F  (5)  Область          — area/region
#   G  (6)  Район            — district
#   H  (7)  Тип міста        — city type abbreviation (м., смт.)
#   I  (8)  Місто (назва)    — city name
#   J  (9)  Тип нас. пункту  — settlement type
#   K  (10) Населений пункт  — settlement name
#   L  (11) Тип вулиці       — street type abbreviation (вул., пров.)
#   M  (12) Вулиця (назва)   — street name
#   N  (13) Будинок          — house number
#   O  (14) Корпус           — building block
#   P  (15) Квартира         — apartment
#   Q  (16) Офіс             — office
#   R  (17) ЄДРПОУ           — EDRPOU tax ID
#   S  (18) ИПН              — individual tax number
#   T  (19) Bank Details     — free-text bank info
#   U–W  (20–22) Contact 1 (Name / Phone / Email)
#   X–Z  (23–25) Contact 2
#   AA–AC (26–28) Contact 3
#   AD–AF (29–31) Contact 4
#   AG–AI (32–34) Contact 5
# One row = one counterparty. No multi-row grouping needed.


def read_sheet(path: str, sheet_name: Optional[str], first_row: int, last_row: int):
    """Yield dicts for each row in the selected sheet (Clients_Flat layout)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row if last_row <= 0 else min(last_row, ws.max_row)
    for idx in range(first_row, max_row + 1):
        row = ws[idx]
        vals = [cell.value for cell in row[:35]]  # A..AI
        vals += [None] * (35 - len(vals))

        # Parse up to 5 contact slots (groups of 3 starting at index 20)
        contacts = []
        for slot in range(5):
            base = 20 + slot * 3  # U=20, X=23, AA=26, AD=29, AG=32
            c_name = safe_str(vals[base])
            c_phone = safe_str(vals[base + 1])
            c_email = safe_str(vals[base + 2])
            if c_name or c_phone or c_email:
                contacts.append({"name": c_name, "phone": c_phone, "email": c_email})

        full_name = safe_str(vals[1])
        short_name = safe_str(vals[2])
        # Fallback: if one is empty, use the other
        if not short_name:
            short_name = full_name
        if not full_name:
            full_name = short_name

        row_dict = {
            "row_idx": idx,
            "manager": safe_str(vals[0]),
            "name": short_name,
            "full_name": full_name,
            "address": safe_str(vals[3]),
            "addr_zip": safe_str(vals[4]),
            "addr_area": safe_str(vals[5]),
            "addr_district": safe_str(vals[6]),
            "addr_city_type": safe_str(vals[7]),
            "addr_city": safe_str(vals[8]),
            "addr_settlement_type": safe_str(vals[9]),
            "addr_settlement": safe_str(vals[10]),
            "addr_street_type": safe_str(vals[11]),
            "addr_street": safe_str(vals[12]),
            "addr_house": safe_str(vals[13]),
            "addr_block": safe_str(vals[14]),
            "addr_apartment": safe_str(vals[15]),
            "addr_office": safe_str(vals[16]),
            "edrpou": safe_digits(vals[17]),
            "ipn": safe_str(vals[18]).replace(" ", ""),
            "bank": safe_str(vals[19]),
            "contacts": contacts,
        }
        yield row_dict


# === BANK DETAILS PARSER ===

IBAN_RE = re.compile(r"UA\d{27}", re.IGNORECASE)
DIGITS_RUN_RE = re.compile(r"\d{10,34}")
MFO_RE = re.compile(r"(?<!\d)\d{6}(?!\d)")
QUOTED_RE = re.compile(r'"([^"]+)"')

BANK_KEYWORDS = (
    "ПРИВАТБАНК", "ОЩАДБАНК", "УКРГАЗБАНК", "РАЙФФАЙЗЕН",
    "УКРСИББАНК", "ПУМБ", "КРЕДОБАНК", "АЛЬФА-БАНК", "ОТП",
)


def parse_bank_details(text: str) -> dict:
    """Extract IBAN / account / MFO / bank name from free-text bank details.
    Returns dict with keys: account, mfo, bank_name (all str, empty if not found)."""
    result = {"account": "", "mfo": "", "bank_name": ""}

    m = IBAN_RE.search(text)
    if m:
        result["account"] = m.group(0).upper()
    else:
        m = DIGITS_RUN_RE.search(text)
        if m:
            result["account"] = m.group(0)

    # MFO: any 6-digit block not inside the already-extracted account
    for m in MFO_RE.finditer(text):
        if result["account"] and m.group(0) in result["account"]:
            continue
        result["mfo"] = m.group(0)
        break

    # Bank name: first quoted substring, or first keyword hit
    qm = QUOTED_RE.search(text)
    if qm:
        result["bank_name"] = qm.group(1)
    else:
        up = text.upper()
        for kw in BANK_KEYWORDS:
            if kw in up:
                result["bank_name"] = kw
                break

    return result


# === PAYLOAD BUILDERS ===

def build_partner_payload(row: dict, manager_key: Optional[str]) -> dict:
    ur_fiz = detect_ur_fiz(row["full_name"])
    payload = {
        "Description": row["name"][:100],
        "DescriptionFull": row["full_name"],
        "ЮрФизЛицо": partner_ur_fiz(ur_fiz),
        "Клиент": True,
        "Конкурент": False,
        "Поставщик": False,
        "ПрочиеОтношения": False,
    }
    if manager_key:
        payload["ОсновнойМенеджер_Key"] = manager_key
    return payload


def build_counterparty_payload(row: dict, partner_key: Optional[str]) -> dict:
    ur_fiz = detect_ur_fiz(row["full_name"])
    is_individual = ur_fiz == "ФизЛицо"
    payload = {
        "Description": row["name"][:100],
        "DescriptionFull": row["full_name"],
        "ЮрФизЛицо": ur_fiz,
    }
    if partner_key:
        payload["Партнер_Key"] = partner_key
    if row["edrpou"] and not is_individual:
        payload["КодПоЕДРПОУ"] = row["edrpou"][:10]
    if row["ipn"]:
        payload["НалоговыйНомер"] = row["ipn"][:50]
        if len(row["ipn"]) <= 12:
            payload["ИННПлательщикаНДС"] = row["ipn"]
    return payload


def build_contact_info_row(line_number: int, ki_type_ref: str, ki_kind: str, representation: str,
                           value_json: str = "") -> dict:
    """Build one row for a ContactInformation tabular section.
    For phone/email `Значение` is left empty (server computes it).
    For addresses, pass `value_json` with the structured JSON string."""
    row = {
        "LineNumber": line_number,
        "Тип": ki_kind,
        "Вид_Key": ki_type_ref,
        "Представление": representation,
        "Значение": value_json,
        "ЗначенияПолей": "",
    }
    if ki_kind == "АдресЭлектроннойПочты":
        row["АдресЭП"] = representation
        if "@" in representation:
            row["ДоменноеИмяСервера"] = representation.split("@", 1)[1]
    return row


def build_address_json(row: dict) -> str:
    """Build a structured address JSON string matching the BSL sibling's ДобавитьАдрес.
    Uses individual fields (city, street, houseNumber, etc.) so that 1C displays
    the structured address form instead of a plain textarea."""
    buildings = []
    if row.get("addr_block"):
        buildings.append({"type": "Корпус", "number": row["addr_block"]})

    apartments = []
    if row.get("addr_apartment"):
        apartments.append({"type": "Квартира", "number": row["addr_apartment"]})
    if row.get("addr_office"):
        apartments.append({"type": "Офис", "number": row["addr_office"]})

    addr = {
        "value":          row.get("address", ""),
        "comment":        "",
        "type":           "Адрес",
        "country":        "Україна",
        "countryCode":    "804",
        "addressType":    "Административно-территориальный",
        "ZIPcode":        row.get("addr_zip", ""),
        "area":           row.get("addr_area", ""),
        "areaType":       "",
        "district":       row.get("addr_district", ""),
        "districtType":   "",
        "city":           row.get("addr_city_type", ""),
        "cityType":       row.get("addr_city", ""),
        "locality":       row.get("addr_settlement_type", ""),
        "localityType":   row.get("addr_settlement", ""),
        "street":         row.get("addr_street_type", ""),
        "streetType":     row.get("addr_street", ""),
        "houseType":      "Дом" if row.get("addr_house") else "",
        "houseNumber":    row.get("addr_house", ""),
        "id": "", "areaCode": "", "areaId": "", "districtId": "",
        "munDistrict": "", "munDistrictType": "", "munDistrictId": "",
        "cityId": "", "cityDistrict": "", "cityDistrictType": "", "cityDistrictId": "",
        "territory": "", "territoryType": "", "territoryId": "",
        "localityId": "", "streetId": "", "houseId": "", "stead": "",
        "buildings":      buildings,
        "apartments":     apartments,
    }
    return json.dumps(addr, ensure_ascii=False)


# === LOOKUPS ===

def find_existing_counterparty(row: dict) -> Optional[dict]:
    """Look up an existing Контрагент by EDRPOU then by name. Returns dict or None."""
    if row["edrpou"]:
        data = odata_get(
            "Catalog_Контрагенты",
            f"$filter=КодПоЕДРПОУ eq '{odata_quote(row['edrpou'])}' and DeletionMark eq false"
            f"&$select=Ref_Key,Партнер_Key&$top=1",
        )
        if data.get("value"):
            return data["value"][0]
    name = row["name"][:100]
    data = odata_get(
        "Catalog_Контрагенты",
        f"$filter=Description eq '{odata_quote(name)}' and DeletionMark eq false"
        f"&$select=Ref_Key,Партнер_Key&$top=1",
    )
    return data["value"][0] if data.get("value") else None


def find_existing_contact_person(partner_key: str, name: str) -> Optional[str]:
    """Return Ref_Key of existing КонтактноеЛицоПартнера by (Владелец, Наименование)."""
    data = odata_get(
        "Catalog_КонтактныеЛицаПартнеров",
        f"$filter=Owner_Key eq guid'{partner_key}' and Description eq '{odata_quote(name[:100])}' "
        f"and DeletionMark eq false&$select=Ref_Key&$top=1",
    )
    items = data.get("value", [])
    return items[0]["Ref_Key"] if items else None


# === WRITERS ===

def create_contact_person(partner_key: str, row: dict, refs: RefCache) -> bool:
    """Create or reuse a КонтактноеЛицоПартнера and attach phone/email.
    Returns True if a new contact was created."""
    name = row["contact_name"] or row["contact_email"] or row["contact_tel"]
    if not name:
        return False

    existing_key = None
    if not DRY_RUN:
        try:
            existing_key = find_existing_contact_person(partner_key, name)
        except Exception:
            existing_key = None

    ci_rows = []
    if row["contact_tel"]:
        phone_ref = refs.ki_type("ТелефонКонтактногоЛица")
        if phone_ref:
            ci_rows.append(build_contact_info_row(
                len(ci_rows) + 1, phone_ref, "Телефон", row["contact_tel"]))
    if row["contact_email"]:
        email_ref = refs.ki_type("EmailКонтактногоЛица")
        if email_ref:
            ci_rows.append(build_contact_info_row(
                len(ci_rows) + 1, email_ref, "АдресЭлектроннойПочты", row["contact_email"]))

    if existing_key:
        odata_patch("Catalog_КонтактныеЛицаПартнеров", existing_key, {
            "ContactInformation": ci_rows,
        })
        return False

    payload = {
        "Owner_Key": partner_key,
        "Description": name[:100],
        "ContactInformation": ci_rows,
    }
    odata_post("Catalog_КонтактныеЛицаПартнеров", payload)
    return True


def create_bank_account(counterparty_key: str, bank_text: str, refs: RefCache) -> bool:
    """Parse free-text bank details and create a БанковскийСчетКонтрагента.
    Returns True if a bank account was created."""
    parsed = parse_bank_details(bank_text)
    if not parsed["account"] and not parsed["mfo"]:
        return False

    payload = {
        "Owner_Key": counterparty_key,
        "Description": (parsed["account"] or bank_text)[:150],
        "НомерСчета": parsed["account"],
        "Комментарий": bank_text,
    }
    if parsed["mfo"]:
        bank_ref = refs.bank_by_mfo(parsed["mfo"])
        if bank_ref:
            payload["Банк_Key"] = bank_ref
        else:
            payload["РучноеИзменениеРеквизитовБанка"] = True
            payload["КодБанка"] = parsed["mfo"]
    if parsed["bank_name"] and "Банк_Key" not in payload:
        payload["РучноеИзменениеРеквизитовБанка"] = True
        payload["НаименованиеБанка"] = parsed["bank_name"][:150]

    try:
        odata_post("Catalog_БанковскиеСчетаКонтрагентов", payload)
        return True
    except Exception as e:
        print(f"  ! bank account write failed: {e}", file=sys.stderr)
        return False


# === MAIN LOOP ===

def _has_address(row: dict) -> bool:
    """Check whether the row contains any address data (presentation or structured)."""
    return bool(row.get("address") or row.get("addr_city") or row.get("addr_street"))


def add_counterparty_contact_info(counterparty_key: str, row: dict, refs: RefCache) -> None:
    """PATCH the Контрагент with ContactInformation rows for the address."""
    if not _has_address(row):
        return
    is_individual = detect_ur_fiz(row["name"]) == "ФизЛицо"
    addr_json = build_address_json(row)
    ci_rows = []
    fact_ref = refs.ki_type("ФактАдресКонтрагента")
    if fact_ref:
        ci_rows.append(build_contact_info_row(
            len(ci_rows) + 1, fact_ref, "Адрес", row["address"], addr_json))
    if not is_individual:
        legal_ref = refs.ki_type("ЮрАдресКонтрагента")
        if legal_ref:
            ci_rows.append(build_contact_info_row(
                len(ci_rows) + 1, legal_ref, "Адрес", row["address"], addr_json))
    if ci_rows:
        odata_patch("Catalog_Контрагенты", counterparty_key,
                    {"ContactInformation": ci_rows})


def add_partner_contact_info(partner_key: str, row: dict, refs: RefCache) -> None:
    """PATCH the Партнер with ContactInformation rows for the address only.
    Phone/email are handled via contact persons (КонтактныеЛицаПартнеров)."""
    ci_rows = []
    if _has_address(row):
        addr_ref = refs.ki_type("АдресПартнера")
        if addr_ref:
            addr_json = build_address_json(row)
            ci_rows.append(build_contact_info_row(
                len(ci_rows) + 1, addr_ref, "Адрес", row["address"], addr_json))
    if ci_rows:
        odata_patch("Catalog_Партнеры", partner_key,
                    {"ContactInformation": ci_rows})


def resolve_manager_key(manager_name: str, cache: dict) -> Optional[str]:
    """Resolve manager name from column A → Пользователи Ref_Key, with caching."""
    if not manager_name:
        return None
    if manager_name in cache:
        return cache[manager_name]
    try:
        data = odata_get(
            "Catalog_Пользователи",
            f"$filter=Description eq '{odata_quote(manager_name)}'&$select=Ref_Key&$top=1",
        )
        items = data.get("value", [])
        key = items[0]["Ref_Key"] if items else None
    except Exception:
        key = None
    if key is None:
        print(f"  ! Manager not found: {manager_name!r}", file=sys.stderr)
    cache[manager_name] = key
    return key


def load_clients(excel_path: str, sheet_name: Optional[str],
                 update: bool, first_row: int, last_row: int) -> dict:
    """Main loader. Mirrors the BSL ЗагрузитьДанныеНаСервере flow."""
    refs = RefCache()
    stats = {"processed": 0, "created": 0, "updated": 0, "contacts": 0, "errors": 0}
    manager_cache: dict[str, Optional[str]] = {}

    for row in read_sheet(excel_path, sheet_name, first_row, last_row):
        if not row["name"]:
            continue

        stats["processed"] += 1

        try:
            # Resolve manager from column A
            mgr_key = resolve_manager_key(row["manager"], manager_cache) if not DRY_RUN else None

            existing = find_existing_counterparty(row) if not DRY_RUN else None

            if existing and not update:
                continue

            if existing:
                partner_key = existing.get("Партнер_Key")
                counterparty_key = existing["Ref_Key"]
                odata_patch("Catalog_Контрагенты", counterparty_key,
                            build_counterparty_payload(row, None))
                stats["updated"] += 1
            else:
                partner = odata_post("Catalog_Партнеры",
                                     build_partner_payload(row, mgr_key))
                partner_key = partner["Ref_Key"]
                add_partner_contact_info(partner_key, row, refs)

                counterparty = odata_post("Catalog_Контрагенты",
                                          build_counterparty_payload(row, partner_key))
                counterparty_key = counterparty["Ref_Key"]
                stats["created"] += 1

            add_counterparty_contact_info(counterparty_key, row, refs)

            if row["bank"]:
                if not create_bank_account(counterparty_key, row["bank"], refs):
                    odata_patch("Catalog_Контрагенты", counterparty_key,
                                {"ДополнительнаяИнформация": row["bank"]})

            # Create contact persons from horizontal slots (up to 5 per row)
            for contact in row["contacts"]:
                fake_row = {"contact_name": contact["name"],
                            "contact_tel": contact["phone"],
                            "contact_email": contact["email"]}
                if create_contact_person(partner_key, fake_row, refs):
                    stats["contacts"] += 1

        except Exception as e:
            stats["errors"] += 1
            print(f"  ! row {row['row_idx']} ({row['name']}) failed: {e}", file=sys.stderr)

    return stats


# === CLI ===

def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--excel", required=True, help="Path to Clients_Flat.xlsx")
    p.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    p.add_argument("--update", action="store_true", help="Update existing counterparties (default: skip)")
    p.add_argument("--first-row", type=int, default=2, help="First data row (default: 2)")
    p.add_argument("--last-row", type=int, default=0, help="Last data row, 0 = all (default: 0)")
    p.add_argument("--dry-run", action="store_true", help="Parse and print only; no OData writes")
    return p.parse_args()


def main():
    global DRY_RUN
    args = parse_args()
    DRY_RUN = args.dry_run

    print(f"Loading {args.excel}" + (f" / sheet={args.sheet}" if args.sheet else ""))
    stats = load_clients(
        args.excel, args.sheet,
        args.update, args.first_row, args.last_row,
    )
    print()
    print("=== RESULT ===")
    print(f"  processed : {stats['processed']}")
    print(f"  created   : {stats['created']}")
    print(f"  updated   : {stats['updated']}")
    print(f"  contacts  : {stats['contacts']}")
    print(f"  errors    : {stats['errors']}")
    sys.exit(1 if stats["errors"] else 0)


if __name__ == "__main__":
    main()
