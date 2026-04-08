#!/usr/bin/env python3
"""
Schüco GPL Sale Price Uploader via OData

Reads a flat Excel (from schueco_flatten_gpl.py) and creates
УстановкаЦенНоменклатуры document(s) in BAS UT via OData.

Flat Excel format: A=Артикул, B=Ціна, C=ОдВим (M/ST), D=ГрупаЗнижки
"""

import argparse
import sys


def upload_prices(excel_path, price_date, price_type_name, batch_size, dry_run):
    """Read flat Excel and upload prices via OData."""
    import openpyxl
    import requests
    import base64

    # OData config (same as nomenclature loader)
    BASE_URL = "http://10.1.5.109/ut_demo/odata/standard.odata"
    _auth_str = base64.b64encode("Адміненко:".encode("utf-8")).decode("ascii")
    HEADERS = {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Basic {_auth_str}",
    }

    def odata_get(entity, params=""):
        url = f"{BASE_URL}/{entity}?{params}&$format=json"
        resp = requests.get(url, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        return resp.json()

    def odata_post(entity, payload):
        url = f"{BASE_URL}/{entity}?$format=json"
        resp = requests.post(url, headers=HEADERS, json=payload, timeout=120)
        if resp.status_code >= 400:
            try:
                msg = resp.json().get("odata.error", {}).get("message", {}).get("value", resp.text)
            except Exception:
                msg = resp.text
            raise Exception(f"OData POST {entity} failed ({resp.status_code}): {msg}")
        return resp.json()

    # Read flat Excel
    print(f"Reading flat Excel: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        art = str(ws.cell(r, 1).value or '').strip()
        price = ws.cell(r, 2).value
        unit = str(ws.cell(r, 3).value or '').strip()
        disc = str(ws.cell(r, 4).value or '').strip()
        if art and price:
            rows.append({'article': art, 'price': float(price), 'unit': unit, 'discount_group': disc})
    wb.close()
    print(f"  Loaded {len(rows)} price rows")

    # Fetch references
    print("Fetching Номенклатура (by Артикул)...")
    nom_map = {}
    skip = 0
    while True:
        data = odata_get("Catalog_Номенклатура",
                         f"$select=Ref_Key,Артикул&$filter=Артикул ne ''&$top=5000&$skip={skip}")
        items = data.get("value", [])
        if not items:
            break
        for item in items:
            art = item["Артикул"].strip().replace(" ", "")
            nom_map[art] = item["Ref_Key"]
        skip += len(items)
    print(f"  Found {len(nom_map)} items with Артикул")

    print("Fetching УпаковкиЕдиницыИзмерения...")
    data = odata_get("Catalog_УпаковкиЕдиницыИзмерения", "$select=Ref_Key,Description")
    uom_map = {}
    for item in data["value"]:
        name = item["Description"].strip().lower()
        uom_map[name] = item["Ref_Key"]
    uom_keys = {}
    for code, aliases in {"M": ["м", "м."], "ST": ["шт", "шт."]}.items():
        for a in aliases:
            if a in uom_map:
                uom_keys[code] = uom_map[a]
                break
    print(f"  UOM keys: {', '.join(f'{k}→{v[:8]}' for k, v in uom_keys.items())}")

    print("Fetching Валюти...")
    data = odata_get("Catalog_Валюты", "$select=Ref_Key,Code")
    eur_key = None
    for item in data["value"]:
        if item["Code"].strip().upper() == "EUR":
            eur_key = item["Ref_Key"]
            break
    if not eur_key:
        print("ERROR: EUR currency not found!")
        return
    print(f"  EUR: {eur_key[:16]}...")

    # Find or create ВидЦены
    print(f"Fetching ВидЦены '{price_type_name}'...")
    data = odata_get("Catalog_ВидыЦен", f"$filter=Description eq '{price_type_name}'&$select=Ref_Key")
    if data.get("value"):
        vid_key = data["value"][0]["Ref_Key"]
        print(f"  Found: {vid_key[:16]}...")
    else:
        if dry_run:
            print(f"  Not found (dry run — won't create)")
            vid_key = None
        else:
            print(f"  Creating '{price_type_name}'...")
            result = odata_post("Catalog_ВидыЦен", {
                "Description": price_type_name,
                "IsFolder": False,
                "ВалютаЦены_Key": eur_key,
                "ЦенаВключаетНДС": False,
                "ИспользоватьПриПродаже": True,
            })
            vid_key = result["Ref_Key"]
            print(f"  Created: {vid_key[:16]}...")

    # Match articles
    matched = []
    unmatched = 0
    for row in rows:
        nom_key = nom_map.get(row['article'])
        if not nom_key:
            unmatched += 1
            continue
        uom_key = uom_keys.get(row['unit'], uom_keys.get('ST'))
        matched.append({
            'nom_key': nom_key,
            'uom_key': uom_key,
            'price': row['price'],
            'article': row['article'],
        })
    print(f"\nMatched: {len(matched)}, Unmatched: {unmatched}")

    if dry_run:
        print("\nDry run — no documents created.")
        print(f"Sample (first 5):")
        for m in matched[:5]:
            print(f"  {m['article']}: {m['price']}")
        return

    if not vid_key:
        print("ERROR: No ВідЦіни key available")
        return

    # Create documents in batches
    total_batches = (len(matched) + batch_size - 1) // batch_size
    for batch_num in range(total_batches):
        start = batch_num * batch_size
        end = min(start + batch_size, len(matched))
        batch = matched[start:end]

        print(f"\nBatch {batch_num + 1}/{total_batches}: {len(batch)} rows...")

        товары = []
        for m in batch:
            товары.append({
                "Номенклатура_Key": m['nom_key'],
                "Характеристика_Key": "00000000-0000-0000-0000-000000000000",
                "Упаковка_Key": m['uom_key'],
                "ВидЦены_Key": vid_key,
                "Цена": m['price'],
            })

        payload = {
            "Date": f"{price_date}T00:00:01",
            "Статус": "Согласован",
            "Комментарий": f"GPL {price_type_name} — автоматичне завантаження (пакет {batch_num + 1})",
            "ВидыЦен": [{"ВидЦены_Key": vid_key}],
            "Товары": товары,
        }

        try:
            result = odata_post("Document_УстановкаЦенНоменклатуры", payload)
            ref_key = result.get("Ref_Key", "?")
            print(f"  Created document: {ref_key[:16]}...")
        except Exception as e:
            print(f"  ERROR: {e}")

    print(f"\nDone. Created {total_batches} document(s) with {len(matched)} price rows.")


def main():
    parser = argparse.ArgumentParser(description="Schüco GPL Sale Price Uploader via OData")
    parser.add_argument("--excel", required=True, help="Path to flat Excel (from schueco_flatten_gpl.py)")
    parser.add_argument("--price-date", required=True, help="Price effective date (YYYY-MM-DD)")
    parser.add_argument("--price-type", default="GPL DE", help="ВідЦіни name (default: GPL DE)")
    parser.add_argument("--batch-size", type=int, default=2000, help="Rows per document (default: 2000)")
    parser.add_argument("--dry-run", action="store_true", help="Match articles, don't create documents")
    args = parser.parse_args()

    upload_prices(args.excel, args.price_date, args.price_type, args.batch_size, args.dry_run)


if __name__ == "__main__":
    main()
