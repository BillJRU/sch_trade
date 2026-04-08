#!/usr/bin/env python3
"""
Schüco GPL (Grundpreisliste) Flattener

Parses the GPL Excel file with side-by-side layout (2-3 articles per row)
and produces a flat Excel with one row per article:
  A=Артикул, B=Ціна, C=ОдВим (M/ST), D=ГрупаЗнижки
"""

import argparse
import re


def parse_article(val):
    """Extract 6-digit article from cell value like '106 116' or 106116."""
    if val is None:
        return None
    s = str(val).strip().replace(" ", "")
    if re.match(r'^\d{5,7}$', s):
        return s
    return None


def extract_from_cells(cells, start, end):
    """Extract price, unit, discount from a slice of cell values."""
    price = None
    unit = None
    discount = None

    for v in cells[start:end]:
        if v is None:
            continue
        sv = str(v).strip()
        if not sv:
            continue

        # Unit detection
        if sv.lower() in ('m', 'meter'):
            unit = 'M'
        elif 'stück' in sv.lower() or 'stuck' in sv.lower() or sv.upper() == 'ST':
            unit = 'ST'
        # Price detection (first float before unit)
        elif unit is None and price is None:
            try:
                fv = float(str(v).replace(',', '.')) if not isinstance(v, (int, float)) else float(v)
                if 0 < fv < 100000:
                    price = fv
            except (ValueError, TypeError):
                pass
        # Discount group (integer 2-3 digits, after unit)
        elif unit is not None and discount is None:
            try:
                iv = int(float(v)) if isinstance(v, (int, float)) else int(float(sv))
                if 10 <= iv <= 999:
                    discount = str(iv)
            except (ValueError, TypeError):
                pass

    return price, unit, discount


def flatten_gpl(excel_path, output_path=None, dry_run=False):
    """Parse GPL Excel → flat list of {article, price, unit, discount_group}."""
    import openpyxl

    print(f"Reading GPL: {excel_path}")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.title}, rows: {ws.max_row}, cols: {ws.max_column}")

    articles = {}
    skipped = 0
    total_found = 0
    row_num = 0

    for row in ws.iter_rows(min_row=34, max_col=65, values_only=True):
        row_num += 1
        if row_num % 1000 == 0:
            print(f"  Processing row {row_num}...")

        cells = list(row)

        # Skip header rows
        is_header = False
        for idx in [0, 3, 4, 7, 10, 33, 35]:
            if idx < len(cells) and cells[idx] and 'Art' in str(cells[idx]):
                is_header = True
                break
        if is_header:
            continue

        # Group 1: article in column A (index 0)
        art1 = parse_article(cells[0]) if cells[0] else None
        if art1:
            price, unit, disc = extract_from_cells(cells, 1, 32)
            if price:
                articles[art1] = {'price': price, 'unit': unit or 'ST', 'discount_group': disc or ''}
                total_found += 1
            else:
                skipped += 1

        # Group 2: article in columns 33-36 (0-based: 32-35)
        art2 = None
        art2_idx = None
        for idx in [32, 33, 34, 35]:
            if idx < len(cells):
                art2 = parse_article(cells[idx])
                if art2:
                    art2_idx = idx
                    break
        if art2 and art2_idx is not None:
            price, unit, disc = extract_from_cells(cells, art2_idx + 1, min(len(cells), 60))
            if price:
                articles[art2] = {'price': price, 'unit': unit or 'ST', 'discount_group': disc or ''}
                total_found += 1
            else:
                skipped += 1

    wb.close()

    # Stats
    unit_dist = {}
    for a in articles.values():
        unit_dist[a['unit']] = unit_dist.get(a['unit'], 0) + 1

    print(f"\nResults:")
    print(f"  Total article-price pairs found: {total_found}")
    print(f"  Unique articles: {len(articles)}")
    print(f"  Skipped (no price): {skipped}")
    print(f"  Unit distribution: {unit_dist}")

    if dry_run or not output_path:
        print(f"\nSample (first 10):")
        for i, (art, data) in enumerate(list(articles.items())[:10]):
            print(f"  {art}: {data['price']} {data['unit']} disc={data['discount_group']}")
        return articles

    # Write flat Excel
    from openpyxl import Workbook
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "GPL Flat"
    out_ws.append(["Артикул", "Ціна", "ОдВим", "ГрупаЗнижки"])
    for art, data in sorted(articles.items()):
        out_ws.append([art, data['price'], data['unit'], data['discount_group']])
    out_wb.save(output_path)
    print(f"\nWritten {len(articles)} rows to: {output_path}")

    return articles


def main():
    parser = argparse.ArgumentParser(description="Schüco GPL Flattener — parse side-by-side layout to flat Excel")
    parser.add_argument("--excel", required=True, help="Path to GPL Excel file")
    parser.add_argument("--output", help="Output flat Excel path (omit for dry-run)")
    parser.add_argument("--dry-run", action="store_true", help="Parse and report, don't write")
    args = parser.parse_args()

    flatten_gpl(args.excel, args.output, args.dry_run)


if __name__ == "__main__":
    main()
